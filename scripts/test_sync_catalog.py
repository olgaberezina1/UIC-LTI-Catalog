import os
import tempfile
import unittest

import sync_catalog as sc

try:
    import openpyxl
except ImportError:  # pragma: no cover - exercised only where openpyxl is absent
    openpyxl = None


class MapStatusTest(unittest.TestCase):
    def test_maps_every_value_the_sheet_currently_uses(self):
        cases = {
            "Yes": "yes",
            "No": "no",
            "NA": "na",
            "Standalone": "standalone",
            "Retired": "retired",
            "Waiting for vendor response": "pending",
            "In Progress": "progress",
            "Available Per Request": "request",
            "Not licensed, unavailable.": "unlicensed",
            "": "—",
            "Status": "—",
        }
        for raw, code in cases.items():
            self.assertEqual(sc.map_status(raw, "Some Tool", "Available in Canvas"), code)

    def test_treats_none_as_em_dash(self):
        self.assertEqual(sc.map_status(None, "Some Tool", "Title II Compliant"), "—")

    def test_strips_surrounding_whitespace(self):
        self.assertEqual(sc.map_status("  Yes  ", "Some Tool", "Available in Canvas"), "yes")

    def test_raises_on_unknown_value_naming_tool_and_column(self):
        with self.assertRaises(sc.SyncError) as ctx:
            sc.map_status("Maybe?", "Packback", "Available in Canvas")
        message = str(ctx.exception)
        self.assertIn("Packback", message)
        self.assertIn("Available in Canvas", message)
        self.assertIn("Maybe?", message)


def row(name="Acadly", bb="Yes", cv="Yes", desc="A tool."):
    return {sc.COL_NAME: name, sc.COL_DESC: desc, sc.COL_BB: bb, sc.COL_CV: cv}


class ShouldKeepTest(unittest.TestCase):
    def test_keeps_a_live_row(self):
        self.assertTrue(sc.should_keep(row()))

    def test_keeps_a_row_live_in_only_one_lms(self):
        self.assertTrue(sc.should_keep(row(bb="Excluded", cv="Yes")))
        self.assertTrue(sc.should_keep(row(bb="Yes", cv="Retired")))

    def test_drops_rows_dead_in_both(self):
        self.assertFalse(sc.should_keep(row(bb="Excluded", cv="Excluded")))
        self.assertFalse(sc.should_keep(row(bb="Retired", cv="Retired")))
        self.assertFalse(sc.should_keep(row(bb="Excluded", cv="Retired")))

    def test_drops_rows_without_a_name(self):
        self.assertFalse(sc.should_keep(row(name="")))
        self.assertFalse(sc.should_keep(row(name="   ")))
        self.assertFalse(sc.should_keep(row(name=None)))

    def test_keeps_a_described_row_with_no_availability_recorded(self):
        # Virtual Machine: a campus computing resource, not an LMS integration,
        # so both availability columns are blank. It is in the catalog today.
        self.assertTrue(sc.should_keep(row(name="Virtual Machine", bb="", cv="")))

    def test_drops_a_placeholder_row_with_no_description(self):
        # Coursera: a bare name with every other cell empty.
        self.assertFalse(sc.should_keep(row(name="Coursera", desc="", bb="", cv="")))


def full_row(**overrides):
    base = {
        sc.COL_NAME: "Panopto",
        sc.COL_DESC: "A comprehensive video platform for education.",
        sc.COL_CATEGORY: "Media & Content Creation",
        sc.COL_BB: "Yes",
        sc.COL_CV: "Yes",
        sc.COL_T2: "Yes",
        sc.COL_VIDEO: None,
        sc.VIDEO_URL_KEY: "",
    }
    base.update(overrides)
    return base


class RowToToolTest(unittest.TestCase):
    def test_maps_the_plain_fields(self):
        self.assertEqual(
            sc.row_to_tool(full_row()),
            {
                "name": "Panopto",
                "desc": "A comprehensive video platform for education.",
                "category": "Media & Content Creation",
                "bb": "yes",
                "cv": "yes",
                "t2": "yes",
            },
        )

    def test_omits_video_keys_when_there_is_no_hyperlink(self):
        tool = sc.row_to_tool(full_row(**{sc.COL_VIDEO: "Watch this"}))
        self.assertNotIn("video", tool)
        self.assertNotIn("videoTitle", tool)

    def test_adds_video_and_title_when_hyperlinked(self):
        tool = sc.row_to_tool(
            full_row(
                **{
                    sc.COL_VIDEO: "How to Embed Panopto Videos",
                    sc.VIDEO_URL_KEY: "https://uic.hosted.panopto.com/Panopto/Pages/Viewer.aspx?id=4006194c",
                }
            )
        )
        self.assertEqual(
            tool["video"],
            "https://uic.hosted.panopto.com/Panopto/Pages/Viewer.aspx?id=4006194c",
        )
        self.assertEqual(tool["videoTitle"], "How to Embed Panopto Videos")

    def test_strips_whitespace_and_tolerates_empty_cells(self):
        tool = sc.row_to_tool(
            full_row(**{sc.COL_NAME: "  Ally  ", sc.COL_DESC: None, sc.COL_T2: ""})
        )
        self.assertEqual(tool["name"], "Ally")
        self.assertEqual(tool["desc"], "")
        self.assertEqual(tool["t2"], "—")

    def test_propagates_a_bad_status_with_the_tool_name(self):
        with self.assertRaises(sc.SyncError) as ctx:
            sc.row_to_tool(full_row(**{sc.COL_NAME: "Labflow", sc.COL_CV: "Sort of"}))
        self.assertIn("Labflow", str(ctx.exception))

    def test_accepts_an_https_video_url(self):
        tool = sc.row_to_tool(
            full_row(
                **{
                    sc.COL_NAME: "Piazza",
                    sc.COL_VIDEO: "Piazza Introduction",
                    sc.VIDEO_URL_KEY: "https://uic.hosted.panopto.com/video1",
                }
            )
        )
        self.assertEqual(tool["video"], "https://uic.hosted.panopto.com/video1")

    def test_rejects_a_plain_http_video_url_naming_the_tool(self):
        with self.assertRaises(sc.SyncError) as ctx:
            sc.row_to_tool(
                full_row(
                    **{
                        sc.COL_NAME: "Piazza",
                        sc.COL_VIDEO: "Piazza Introduction",
                        sc.VIDEO_URL_KEY: "http://uic.hosted.panopto.com/video1",
                    }
                )
            )
        self.assertIn("Piazza", str(ctx.exception))

    def test_rejects_a_javascript_video_url_naming_the_tool(self):
        with self.assertRaises(sc.SyncError) as ctx:
            sc.row_to_tool(
                full_row(
                    **{
                        sc.COL_NAME: "Piazza",
                        sc.COL_VIDEO: "Piazza Introduction",
                        sc.VIDEO_URL_KEY: "javascript:alert(1)",
                    }
                )
            )
        self.assertIn("Piazza", str(ctx.exception))


class BuildCatalogTest(unittest.TestCase):
    def many_rows(self, count, prefix="Tool"):
        return [full_row(**{sc.COL_NAME: f"{prefix} {i:03d}"}) for i in range(count)]

    def test_drops_unpublished_rows(self):
        rows = self.many_rows(sc.MIN_TOOLS)
        rows.append(full_row(**{sc.COL_NAME: "Kortex", sc.COL_BB: "Retired", sc.COL_CV: "Retired"}))
        rows.append(full_row(**{sc.COL_NAME: "", sc.COL_BB: "Yes", sc.COL_CV: "Yes"}))
        tools, _ = sc.build_catalog(rows)
        names = [t["name"] for t in tools]
        self.assertEqual(len(names), sc.MIN_TOOLS)
        self.assertNotIn("Kortex", names)

    def test_keeps_the_first_of_a_duplicate_and_warns(self):
        rows = self.many_rows(sc.MIN_TOOLS)
        rows.append(full_row(**{sc.COL_NAME: "Packback", sc.COL_CV: "Yes"}))
        rows.append(full_row(**{sc.COL_NAME: "Packback", sc.COL_CV: "No"}))
        tools, warnings = sc.build_catalog(rows)
        packback = [t for t in tools if t["name"] == "Packback"]
        self.assertEqual(len(packback), 1)
        self.assertEqual(packback[0]["cv"], "yes")
        self.assertEqual(len(warnings), 1)
        self.assertIn("Packback", warnings[0])

    def test_sorts_case_insensitively_by_name(self):
        rows = self.many_rows(sc.MIN_TOOLS, prefix="Zed")
        rows.append(full_row(**{sc.COL_NAME: "ACS Lab Safety UIC"}))
        rows.append(full_row(**{sc.COL_NAME: "Acadly"}))
        rows.append(full_row(**{sc.COL_NAME: "ATI Testing"}))
        tools, _ = sc.build_catalog(rows)
        self.assertEqual(
            [t["name"] for t in tools[:3]],
            ["Acadly", "ACS Lab Safety UIC", "ATI Testing"],
        )

    def test_refuses_to_build_a_suspiciously_small_catalog(self):
        with self.assertRaises(sc.SyncError) as ctx:
            sc.build_catalog(self.many_rows(sc.MIN_TOOLS - 1))
        self.assertIn(str(sc.MIN_TOOLS), str(ctx.exception))

    def test_warns_when_description_contains_a_newline(self):
        rows = self.many_rows(sc.MIN_TOOLS - 1)
        rows.append(
            full_row(
                **{
                    sc.COL_NAME: "Gradescope",
                    sc.COL_DESC: "An AI-assisted grading platform.\n\nGradescope Linking Assignments",
                }
            )
        )
        tools, warnings = sc.build_catalog(rows)
        self.assertEqual(len(tools), sc.MIN_TOOLS)
        self.assertEqual(len(warnings), 1)
        self.assertIn("Gradescope", warnings[0])
        self.assertIn("Gradescope Linking Assignments", warnings[0])
        self.assertIn("clean the sheet's Description cell", warnings[0])
        # The sheet is the source of truth: the description itself is untouched.
        kept = [t for t in tools if t["name"] == "Gradescope"][0]
        self.assertEqual(
            kept["desc"], "An AI-assisted grading platform.\n\nGradescope Linking Assignments"
        )

    def test_warns_when_last_line_matches_video_title_even_without_a_newline(self):
        rows = self.many_rows(sc.MIN_TOOLS - 1)
        rows.append(
            full_row(
                **{
                    sc.COL_NAME: "Piazza",
                    sc.COL_DESC: "Piazza Introduction",
                    sc.COL_VIDEO: "Piazza Introduction",
                    sc.VIDEO_URL_KEY: "https://uic.hosted.panopto.com/video1",
                }
            )
        )
        tools, warnings = sc.build_catalog(rows)
        self.assertEqual(len(warnings), 1)
        self.assertIn("Piazza", warnings[0])

    def test_no_warning_for_a_clean_description(self):
        rows = self.many_rows(sc.MIN_TOOLS - 1)
        rows.append(full_row(**{sc.COL_NAME: "Acadly", sc.COL_DESC: "A clean, single-line tool."}))
        tools, warnings = sc.build_catalog(rows)
        self.assertEqual(warnings, [])


def tool(name="Acadly", **overrides):
    base = {"name": name, "desc": "d", "category": "c", "bb": "yes", "cv": "yes", "t2": "yes"}
    base.update(overrides)
    return base


class DiffCatalogsTest(unittest.TestCase):
    def test_reports_nothing_when_identical(self):
        self.assertEqual(sc.diff_catalogs([tool()], [tool()]), [])

    def test_reports_additions_and_removals(self):
        lines = sc.diff_catalogs([tool("Kortex")], [tool("ATI Testing")])
        self.assertIn("+ ATI Testing", lines)
        self.assertIn("- Kortex", lines)

    def test_reports_a_changed_status_with_both_values(self):
        lines = sc.diff_catalogs(
            [tool("ClassRanked", bb="no", cv="contract")],
            [tool("ClassRanked", bb="yes", cv="yes")],
        )
        self.assertEqual(lines, ["ClassRanked: bb no->yes, cv contract->yes"])

    def test_summarises_description_changes_without_quoting_them(self):
        lines = sc.diff_catalogs(
            [tool("Gradescope", desc="An AI-assisted grading platform. Gradescope Linking Assignments")],
            [tool("Gradescope", desc="An AI-assisted grading platform.")],
        )
        self.assertEqual(lines, ["Gradescope: desc changed"])

    def test_reports_a_newly_added_video(self):
        lines = sc.diff_catalogs(
            [tool("Panopto")],
            [tool("Panopto", video="https://example.test/v", videoTitle="Intro")],
        )
        self.assertEqual(len(lines), 1)
        self.assertIn("video", lines[0])


REQUIRED_HEADERS = [
    sc.COL_NAME,
    sc.COL_DESC,
    sc.COL_CATEGORY,
    sc.COL_BB,
    sc.COL_CV,
    sc.COL_T2,
    sc.COL_VIDEO,
]


@unittest.skipUnless(openpyxl, "openpyxl is not installed")
class ReadRowsTest(unittest.TestCase):
    def make_book(self, path, sheet_title, headers, rows=(), hyperlinks=None):
        """Save a workbook to `path`. `hyperlinks` maps (row, col) -> url, both
        1-based and counted from the header row (so row 2 is the first data row).
        """
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = sheet_title
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        for r, values in enumerate(rows, start=2):
            for col, value in enumerate(values, start=1):
                sheet.cell(row=r, column=col, value=value)
        for (r, col), url in (hyperlinks or {}).items():
            sheet.cell(row=r, column=col).hyperlink = url
        book.save(path)

    def test_reads_rows_keyed_by_header_skips_blank_and_carries_hyperlink(self):
        headers = REQUIRED_HEADERS + [None]  # one blank-header column, to be skipped
        rows = [
            [
                "Piazza",
                "A discussion tool.",
                "Collaboration & Communication",
                "Yes",
                "Yes",
                "Yes",
                "Piazza Introduction",
                "ignored",
            ],
            [
                "Zoom",
                "Video conferencing.",
                "Media & Content Creation",
                "Yes",
                "Yes",
                "No",
                "Zoom Basics",
                "ignored",
            ],
        ]
        video_col = headers.index(sc.COL_VIDEO) + 1
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "book.xlsx")
            self.make_book(
                path,
                sc.SHEET_TAB,
                headers,
                rows,
                hyperlinks={(2, video_col): "https://example.test/piazza-video"},
            )
            result = sc.read_rows(path)

        self.assertEqual(len(result), 2)
        self.assertNotIn(None, result[0])
        self.assertEqual(result[0][sc.COL_NAME], "Piazza")
        self.assertEqual(result[0][sc.COL_DESC], "A discussion tool.")
        self.assertEqual(result[0][sc.VIDEO_URL_KEY], "https://example.test/piazza-video")
        self.assertEqual(result[1][sc.COL_NAME], "Zoom")
        self.assertEqual(result[1][sc.VIDEO_URL_KEY], "")

    def test_raises_when_the_sheet_tab_is_missing(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "book.xlsx")
            self.make_book(path, "Some Other Tab", REQUIRED_HEADERS)
            with self.assertRaises(sc.SyncError):
                sc.read_rows(path)

    def test_raises_naming_the_missing_description_column(self):
        headers = [h for h in REQUIRED_HEADERS if h != sc.COL_DESC]
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "book.xlsx")
            self.make_book(path, sc.SHEET_TAB, headers)
            with self.assertRaises(sc.SyncError) as ctx:
                sc.read_rows(path)
            self.assertIn(sc.COL_DESC, str(ctx.exception))


if __name__ == "__main__":
    unittest.main()
