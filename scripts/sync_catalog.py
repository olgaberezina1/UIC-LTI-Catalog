#!/usr/bin/env python3
"""Sync the UIC LTI Catalog from its Google Sheet into tools.json."""

import argparse
import json
import os
import subprocess
import sys
import tempfile
import urllib.error
import urllib.request
import zipfile

SHEET_ID = "1FGY1itGZgsnpefzKDXtfqH2AZdkVCYlnQxt_IInFqXY"
SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
SHEET_TAB = "Published"

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TOOLS_JSON = os.path.join(REPO_ROOT, "tools.json")

MIN_TOOLS = 40

RECOVERY_HINT = (
    "tools.json is written locally — inspect with 'git status', then "
    "'git push origin main' once it is committed"
)

COL_NAME = "Tool Name1"
COL_DESC = "Description"
COL_CATEGORY = "Category"
COL_BB = "Available in Blackboard"
COL_CV = "Available in Canvas"
COL_T2 = "Title II Compliant"
COL_VIDEO = "UIC Walkthrough Video"

# read_rows() stashes each walkthrough cell's hyperlink target under this key,
# because the visible cell text is a title like "Piazza Introduction".
VIDEO_URL_KEY = "_video_url"

# Sheet wording -> the status codes AVAIL_LABEL already renders (app.js:66).
STATUS_MAP = {
    "Yes": "yes",
    "No": "no",
    "NA": "na",
    "Standalone": "standalone",
    "Retired": "retired",
    "Waiting for vendor response": "pending",
    "In Progress": "progress",
    "Available Per Request": "request",
    "Not licensed, unavailable.": "unlicensed",
    "": "—",
    "Status": "—",
}

# A row dead in both columns is not published. "Excluded" is the sheet's own
# opt-out flag. Blank is deliberately NOT here: Virtual Machine is blank in both
# columns because it is a campus computing resource rather than an LMS
# integration, and it belongs in the catalog. Content-free rows are caught by the
# description check in should_keep() instead.
DROP_STATUSES = {"Excluded", "Retired"}


class SyncError(Exception):
    """Raised when the sheet holds something the script must not guess about."""


def cell(row, key):
    """A sheet cell as a stripped string. Cells come back as None when empty."""
    return str(row.get(key) or "").strip()


def map_status(value, tool, column):
    """Turn one availability cell into its status code."""
    key = str(value or "").strip()
    if key not in STATUS_MAP:
        raise SyncError(
            f"{tool}: unrecognized {column} value {key!r}. "
            f"Fix the sheet, or add the value to STATUS_MAP "
            f"(and to AVAIL_LABEL in app.js if it needs a new label)."
        )
    return STATUS_MAP[key]


def should_keep(row):
    """True when a row belongs in the published catalog."""
    if not cell(row, COL_NAME):
        return False
    # A name with nothing behind it is a placeholder, not a tool.
    if not cell(row, COL_DESC):
        return False
    return not (
        cell(row, COL_BB) in DROP_STATUSES and cell(row, COL_CV) in DROP_STATUSES
    )


def row_to_tool(row):
    """Convert one kept sheet row into a catalog entry."""
    name = cell(row, COL_NAME)
    tool = {
        "name": name,
        "desc": cell(row, COL_DESC),
        "category": cell(row, COL_CATEGORY),
        "bb": map_status(row.get(COL_BB), name, COL_BB),
        "cv": map_status(row.get(COL_CV), name, COL_CV),
        "t2": map_status(row.get(COL_T2), name, COL_T2),
    }
    url = cell(row, VIDEO_URL_KEY)
    if url:
        if not url.startswith("https://"):
            raise SyncError(
                f"{name}: walkthrough video URL {url!r} is not https — fix the "
                f"sheet's hyperlink"
            )
        tool["video"] = url
        tool["videoTitle"] = cell(row, COL_VIDEO)
    return tool


def lint_description(tool):
    """A non-fatal note when a Description cell needs sheet cleanup, else None.

    The sheet is the source of truth and this script never edits content — it
    only flags descriptions that carry a newline, or whose trailing line
    duplicates the walkthrough video's title (a leftover from pasting the
    video's link text into the Description cell).
    """
    desc = tool["desc"]
    lines = [line.strip() for line in desc.split("\n") if line.strip()]
    last_line = lines[-1] if lines else ""
    video_title = tool.get("videoTitle", "").strip()
    trailing_matches_title = bool(video_title) and last_line.lower() == video_title.lower()

    if "\n" not in desc and not trailing_matches_title:
        return None

    return (
        f"{tool['name']}: description carries a trailing line {last_line!r} — "
        f"clean the sheet's Description cell"
    )


def build_catalog(rows):
    """Filter, convert, de-duplicate and sort every sheet row.

    Returns (tools, warnings). Warnings are non-fatal notes for the operator.
    """
    tools = []
    warnings = []
    seen = set()

    for row in rows:
        if not should_keep(row):
            continue
        tool = row_to_tool(row)
        if tool["name"] in seen:
            warnings.append(f"duplicate row for {tool['name']} — keeping the first")
            continue
        seen.add(tool["name"])
        tools.append(tool)
        lint = lint_description(tool)
        if lint:
            warnings.append(lint)

    tools.sort(key=lambda t: t["name"].lower())

    if len(tools) < MIN_TOOLS:
        raise SyncError(
            f"only {len(tools)} tools survived filtering, minimum is {MIN_TOOLS}. "
            f"The {SHEET_TAB!r} tab may have been renamed, restructured or unshared. "
            f"Refusing to empty the catalog."
        )
    return tools, warnings


DIFF_FIELDS = ("desc", "category", "bb", "cv", "t2", "video", "videoTitle")


def diff_catalogs(old, new):
    """One readable line per difference, for the console and the commit message."""
    old_by = {t["name"]: t for t in old}
    new_by = {t["name"]: t for t in new}
    lines = []

    for name in sorted(set(new_by) - set(old_by), key=str.lower):
        lines.append(f"+ {name}")
    for name in sorted(set(old_by) - set(new_by), key=str.lower):
        lines.append(f"- {name}")

    for name in sorted(set(old_by) & set(new_by), key=str.lower):
        changes = []
        for field in DIFF_FIELDS:
            before = old_by[name].get(field, "")
            after = new_by[name].get(field, "")
            if before == after:
                continue
            if field == "desc":
                changes.append("desc changed")
            else:
                changes.append(f"{field} {before or '(none)'}->{after or '(none)'}")
        if changes:
            lines.append(f"{name}: " + ", ".join(changes))

    return lines


def download_xlsx(dest):
    """Fetch the sheet as xlsx. Only this format preserves hyperlink targets."""
    with urllib.request.urlopen(SHEET_URL, timeout=60) as response:
        payload = response.read()

    # A shared sheet returns a zip container; an unshared one returns a
    # sign-in page with a cheerful 200.
    if not payload.startswith(b"PK"):
        raise SyncError(
            "the sheet did not return a spreadsheet. Check that it is still "
            "shared as 'anyone with the link can view'."
        )

    with open(dest, "wb") as handle:
        handle.write(payload)


def read_rows(path):
    """Read the Published tab into dicts keyed by header, plus the video URL."""
    import openpyxl  # imported here so the pure-function tests need not install it

    workbook = openpyxl.load_workbook(path, data_only=True)
    if SHEET_TAB not in workbook.sheetnames:
        raise SyncError(
            f"no {SHEET_TAB!r} tab in the workbook, found {workbook.sheetnames}"
        )

    sheet = workbook[SHEET_TAB]
    headers = [c.value for c in sheet[1]]
    for required in (COL_NAME, COL_DESC, COL_BB, COL_CV, COL_T2, COL_CATEGORY, COL_VIDEO):
        if required not in headers:
            raise SyncError(f"expected a {required!r} column, found {headers}")

    video_column = headers.index(COL_VIDEO) + 1
    rows = []
    for number in range(2, sheet.max_row + 1):
        row = {
            header: sheet.cell(row=number, column=index + 1).value
            for index, header in enumerate(headers)
            if header
        }
        link = sheet.cell(row=number, column=video_column).hyperlink
        row[VIDEO_URL_KEY] = link.target if link else ""
        rows.append(row)
    return rows


def git(*args):
    """Run one git command in the repo, raising if it fails."""
    result = subprocess.run(
        ("git",) + args,
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise SyncError(f"git {' '.join(args)} failed:\n{result.stderr.strip()}")
    return result.stdout


def is_ancestor(ancestor, ref):
    """True when `ancestor` is reachable from `ref` (git merge-base --is-ancestor).

    Not routed through git() above: a "no" answer is a normal, non-zero exit
    for this command, not a failure to report as one.
    """
    result = subprocess.run(
        ("git", "merge-base", "--is-ancestor", ancestor, ref),
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
    )
    return result.returncode == 0


def ensure_publishable():
    """Pre-flight for the real (non-dry-run) path, before write_tools() runs.

    The script publishes unattended, so it refuses to run anywhere but a
    local main that is caught up with origin — otherwise a sync from a stale
    checkout could push a confusing merge or silently clobber newer history.
    """
    branch = git("rev-parse", "--abbrev-ref", "HEAD").strip()
    if branch != "main":
        raise SyncError(
            f"on branch '{branch}' — the script publishes from main; switch branches first"
        )
    git("fetch", "origin", "main")
    if not is_ancestor("origin/main", "HEAD"):
        raise SyncError("local main is behind origin/main — pull first")


def commit_and_push(summary):
    """Commit tools.json alone, then push. Never stages anything else."""
    git("add", "--", TOOLS_JSON)
    message = (
        "Sync catalog from Google Sheet\n\n"
        + "\n".join(summary)
        + "\n\nGenerated-By: scripts/sync_catalog.py\n"
    )
    git("commit", "-m", message)
    git("push", "origin", "main")


def load_existing():
    if not os.path.exists(TOOLS_JSON):
        return []
    with open(TOOLS_JSON, encoding="utf-8") as handle:
        return json.load(handle)


def write_tools(tools):
    with open(TOOLS_JSON, "w", encoding="utf-8") as handle:
        json.dump(tools, handle, indent=2, ensure_ascii=False)
        handle.write("\n")


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="print the change summary without writing, committing or pushing",
    )
    args = parser.parse_args(argv)

    with tempfile.TemporaryDirectory() as workspace:
        xlsx = os.path.join(workspace, "catalog.xlsx")
        download_xlsx(xlsx)
        rows = read_rows(xlsx)

    tools, warnings = build_catalog(rows)
    for warning in warnings:
        print(f"warning: {warning}")

    summary = diff_catalogs(load_existing(), tools)
    print(f"{len(tools)} tools in the sheet")

    if not summary:
        print("no changes")
        return 0

    for line in summary:
        print(f"  {line}")

    if args.dry_run:
        print("\ndry run — nothing written")
        return 0

    ensure_publishable()

    write_tools(tools)
    try:
        commit_and_push(summary)
    except Exception:
        # Whatever went wrong here (SyncError from git(), or anything else),
        # tools.json is already on disk — say so before the exception propagates.
        print(RECOVERY_HINT, file=sys.stderr)
        raise
    print(f"\ncommitted and pushed {len(summary)} change(s)")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except (
        SyncError,
        urllib.error.URLError,
        zipfile.BadZipFile,
        OSError,
        subprocess.CalledProcessError,
    ) as error:
        # Known failure modes print in the script's own format. Anything else
        # is unrecognized and should traceback loudly rather than be muffled.
        print(f"error: {error}", file=sys.stderr)
        sys.exit(1)
